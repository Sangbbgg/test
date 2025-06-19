import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import os
from datetime import datetime
import csv

# pandas 대신 openpyxl만 사용하여 의존성 문제 해결
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelCreator:
    def __init__(self, root):
        self.root = root
        self.root.title("엑셀 파일 생성기")
        self.root.geometry("600x500")
        
        # 데이터 저장용
        self.data = []
        
        self.setup_ui()
    
    def setup_ui(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 제목
        title_label = ttk.Label(main_frame, text="엑셀 파일 생성기", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # 데이터 입력 섹션
        input_frame = ttk.LabelFrame(main_frame, text="데이터 입력", padding="10")
        input_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # 이름 입력
        ttk.Label(input_frame, text="이름:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.name_entry = ttk.Entry(input_frame, width=20)
        self.name_entry.grid(row=0, column=1, padx=5)
        
        # 나이 입력
        ttk.Label(input_frame, text="나이:").grid(row=0, column=2, sticky=tk.W, padx=(10, 5))
        self.age_entry = ttk.Entry(input_frame, width=10)
        self.age_entry.grid(row=0, column=3, padx=5)
        
        # 이메일 입력
        ttk.Label(input_frame, text="이메일:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(10, 0))
        self.email_entry = ttk.Entry(input_frame, width=30)
        self.email_entry.grid(row=1, column=1, columnspan=2, padx=5, pady=(10, 0), sticky=(tk.W, tk.E))
        
        # 부서 입력
        ttk.Label(input_frame, text="부서:").grid(row=2, column=0, sticky=tk.W, padx=(0, 5), pady=(10, 0))
        self.dept_entry = ttk.Entry(input_frame, width=20)
        self.dept_entry.grid(row=2, column=1, padx=5, pady=(10, 0))
        
        # 추가 버튼
        add_btn = ttk.Button(input_frame, text="데이터 추가", command=self.add_data)
        add_btn.grid(row=2, column=3, padx=5, pady=(10, 0))
        
        # 데이터 목록 표시
        list_frame = ttk.LabelFrame(main_frame, text="입력된 데이터", padding="10")
        list_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        
        # 트리뷰 (표 형태로 데이터 표시)
        columns = ("이름", "나이", "이메일", "부서")
        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=8)
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120)
        
        self.tree.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.grid(row=0, column=2, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # 삭제 버튼
        delete_btn = ttk.Button(list_frame, text="선택 삭제", command=self.delete_selected)
        delete_btn.grid(row=1, column=0, pady=(10, 0), sticky=tk.W)
        
        # 모두 삭제 버튼
        clear_btn = ttk.Button(list_frame, text="모두 삭제", command=self.clear_all)
        clear_btn.grid(row=1, column=1, pady=(10, 0), sticky=tk.W)
        
        # 파일 생성 섹션
        file_frame = ttk.LabelFrame(main_frame, text="파일 생성", padding="10")
        file_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)
        
        # 파일명 입력
        ttk.Label(file_frame, text="파일명:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.filename_entry = ttk.Entry(file_frame, width=30)
        self.filename_entry.grid(row=0, column=1, padx=5)
        self.filename_entry.insert(0, f"직원명단_{datetime.now().strftime('%Y%m%d')}")
        
        # 샘플 데이터 버튼
        sample_btn = ttk.Button(file_frame, text="샘플 데이터", command=self.add_sample_data)
        sample_btn.grid(row=0, column=2, padx=5)
        
        # 파일 형식 선택
        format_frame = ttk.Frame(file_frame)
        format_frame.grid(row=1, column=0, columnspan=3, pady=(5, 0))
        
        ttk.Label(format_frame, text="파일 형식:").grid(row=0, column=0, sticky=tk.W)
        self.format_var = tk.StringVar(value="xlsx")
        ttk.Radiobutton(format_frame, text="Excel (.xlsx)", variable=self.format_var, value="xlsx").grid(row=0, column=1, padx=5)
        ttk.Radiobutton(format_frame, text="CSV (.csv)", variable=self.format_var, value="csv").grid(row=0, column=2, padx=5)
        
        # 파일 생성 버튼
        create_btn = ttk.Button(file_frame, text="파일 생성", command=self.create_file)
        create_btn.grid(row=2, column=0, columnspan=3, pady=(10, 0))
        
        # 그리드 가중치 설정
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(0, weight=1)
    
    def add_data(self):
        name = self.name_entry.get().strip()
        age = self.age_entry.get().strip()
        email = self.email_entry.get().strip()
        dept = self.dept_entry.get().strip()
        
        if not name:
            messagebox.showwarning("경고", "이름을 입력해주세요.")
            return
        
        try:
            age_int = int(age) if age else 0
        except ValueError:
            messagebox.showwarning("경고", "나이는 숫자로 입력해주세요.")
            return
        
        # 데이터 추가
        data_row = [name, age_int, email, dept]
        self.data.append(data_row)
        
        # 트리뷰에 추가
        self.tree.insert("", "end", values=data_row)
        
        # 입력 필드 초기화
        self.name_entry.delete(0, tk.END)
        self.age_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.dept_entry.delete(0, tk.END)
        
        self.name_entry.focus()
    
    def delete_selected(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("경고", "삭제할 항목을 선택해주세요.")
            return
        
        # 선택된 항목들을 역순으로 삭제 (인덱스 변경 문제 방지)
        items_to_delete = []
        for item in selected:
            index = self.tree.index(item)
            items_to_delete.append((index, item))
        
        # 인덱스 기준으로 역순 정렬
        items_to_delete.sort(key=lambda x: x[0], reverse=True)
        
        for index, item in items_to_delete:
            self.tree.delete(item)
            del self.data[index]
    
    def clear_all(self):
        if messagebox.askyesno("확인", "모든 데이터를 삭제하시겠습니까?"):
            self.tree.delete(*self.tree.get_children())
            self.data.clear()
    
    def add_sample_data(self):
        sample_data = [
            ["김철수", 30, "kim@company.com", "개발팀"],
            ["이영희", 28, "lee@company.com", "디자인팀"],
            ["박민수", 35, "park@company.com", "마케팅팀"],
            ["최지은", 26, "choi@company.com", "인사팀"],
            ["정호영", 32, "jung@company.com", "개발팀"]
        ]
        
        for data_row in sample_data:
            self.data.append(data_row)
            self.tree.insert("", "end", values=data_row)
    
    def create_file(self):
        if not self.data:
            messagebox.showwarning("경고", "생성할 데이터가 없습니다.")
            return
        
        filename = self.filename_entry.get().strip()
        if not filename:
            filename = f"직원명단_{datetime.now().strftime('%Y%m%d')}"
        
        file_format = self.format_var.get()
        
        if file_format == "xlsx":
            self.create_excel_file(filename)
        else:
            self.create_csv_file(filename)
    
    def create_excel_file(self, filename):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("오류", "Excel 파일 생성에 필요한 라이브러리가 없습니다.\nCSV 형식을 사용해주세요.")
            return
        
        # 확장자 추가
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        try:
            # 저장 위치 선택
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialvalue=filename
            )
            
            if not file_path:
                return
            
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "직원명단"
            
            # 헤더 추가
            headers = ["이름", "나이", "이메일", "부서"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # 데이터 추가
            for row, data_row in enumerate(self.data, 2):
                for col, value in enumerate(data_row, 1):
                    ws.cell(row=row, column=col, value=value)
            
            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 최대 너비 제한
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(file_path)
            messagebox.showinfo("성공", f"Excel 파일이 생성되었습니다!\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 생성 중 오류가 발생했습니다:\n{str(e)}")
    
    def create_csv_file(self, filename):
        # 확장자 추가
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        try:
            # 저장 위치 선택
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialvalue=filename
            )
            
            if not file_path:
                return
            
            # CSV 파일 생성
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                
                # 헤더 작성
                writer.writerow(["이름", "나이", "이메일", "부서"])
                
                # 데이터 작성
                writer.writerows(self.data)
            
            messagebox.showinfo("성공", f"CSV 파일이 생성되었습니다!\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("오류", f"파일 생성 중 오류가 발생했습니다:\n{str(e)}")

def main():
    root = tk.Tk()
    app = ExcelCreator(root)
    
    # Enter 키로 데이터 추가
    root.bind('<Return>', lambda e: app.add_data())
    
    root.mainloop()

if __name__ == "__main__":
    main()